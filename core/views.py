from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, F
from django.http import JsonResponse, HttpResponse
from django.forms import modelformset_factory
from django.db import transaction
from django.template.loader import get_template
from django.contrib.auth.decorators import user_passes_test
from xhtml2pdf import pisa
from openpyxl import Workbook
from io import BytesIO
from .models import Product, Order, OrderItem, StockMovement
from .forms import ProductForm, OrderForm, OrderItemForm, StockMovementForm
from accounts.models import User

def home(request):
    if request.user.is_authenticated:
        return redirect('core:dashboard')
    return render(request, 'core/home.html')

def dashboard(request):
    # Statistiques
    total_products = Product.objects.count()
    total_orders = Order.objects.count()
    total_stock_movements = StockMovement.objects.count()
    
    # Dernières commandes
    recent_orders = Order.objects.order_by('-created_at')[:5]
    
    # Derniers mouvements de stock
    recent_movements = StockMovement.objects.order_by('-created_at')[:5]
    
    context = {
        'total_products': total_products,
        'total_orders': total_orders,
        'total_stock_movements': total_stock_movements,
        'recent_orders': recent_orders,
        'recent_movements': recent_movements
    }
    return render(request, 'core/dashboard.html', context)

# Vues d'export pour l'historique du stock
@login_required
def export_stock_pdf(request):
    movements = StockMovement.objects.all()
    
    # Rendre le template en PDF
    template = get_template('core/stock_pdf.html')
    context = {
        'movements': movements,
        'total_in': movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0,
        'total_out': movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
    }
    html = template.render(context)
    
    # Créer le PDF
    pdf = BytesIO()
    pisa.CreatePDF(html, dest=pdf)
    
    # Configurer la réponse
    response = HttpResponse(pdf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=historique_stock.pdf'
    
    return response

@user_passes_test(lambda u: u.is_admin_user())
@login_required
def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)
    OrderItemFormSet = modelformset_factory(OrderItem, form=OrderItemForm, extra=1)
    
    if request.method == 'POST':
        formset = OrderItemFormSet(request.POST, queryset=OrderItem.objects.filter(order=order))
        if formset.is_valid():
            with transaction.atomic():
                # Delete removed items
                for form in formset.deleted_forms:
                    if form.cleaned_data.get('DELETE'):
                        item = form.save(commit=False)
                        item.product.quantity += item.quantity
                        item.product.save()
                        item.delete()
                
                # Update existing items and create new ones
                for form in formset:
                    if form.cleaned_data and not form.cleaned_data.get('DELETE'):
                        item = form.save(commit=False)
                        if item.pk:
                            # Update existing item
                            old_item = OrderItem.objects.get(pk=item.pk)
                            item.product.quantity += old_item.quantity
                            item.product.quantity -= item.quantity
                        else:
                            # New item
                            item.product.quantity -= item.quantity
                            item.order = order
                        item.product.save()
                        item.save()
                
                # Update order status if all items are deleted
                if not order.items.exists():
                    order.status = 'draft'
                order.save()
                
                messages.success(request, 'Commande mise à jour avec succès')
                return redirect('core:order_list')
    else:
        formset = OrderItemFormSet(queryset=OrderItem.objects.filter(order=order))
    
    return render(request, 'core/order_edit.html', {
        'order': order,
        'formset': formset
    })

@user_passes_test(lambda u: u.is_admin_user())
@login_required
def order_delete(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        with transaction.atomic():
            # Return stock quantities
            for item in order.items.all():
                item.product.quantity += item.quantity
                item.product.save()
            order.delete()
            messages.success(request, 'Commande supprimée avec succès')
        return redirect('core:order_list')
    return render(request, 'core/order_confirm_delete.html', {'order': order})

@login_required
def export_stock_excel(request):
    movements = StockMovement.objects.all()
    
    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique Stock"
    
    # Ajouter les en-têtes
    headers = ['Date', 'Produit', 'Type', 'Quantité', 'Motif', 'Source', 'Destination']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Ajouter les mouvements
    row_num = 2
    for movement in movements:
        ws.cell(row=row_num, column=1, value=movement.created_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=2, value=movement.product.name)
        ws.cell(row=row_num, column=3, value=movement.get_movement_type_display())
        ws.cell(row=row_num, column=4, value=movement.quantity)
        ws.cell(row=row_num, column=5, value=movement.reason)
        ws.cell(row=row_num, column=6, value=movement.source)
        ws.cell(row=row_num, column=7, value=movement.destination)
        row_num += 1
    
    # Ajouter les totaux
    ws.cell(row=row_num, column=3, value='Total Entrées:')
    ws.cell(row=row_num, column=4, value=movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0)
    row_num += 1
    ws.cell(row=row_num, column=3, value='Total Sorties:')
    ws.cell(row=row_num, column=4, value=movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0)
    
    # Sauvegarder le workbook
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Configurer la réponse
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=historique_stock.xlsx'
    
    return response

# Vues pour la gestion des produits
@user_passes_test(lambda u: u.is_admin_user())
@login_required
def product_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()
            messages.success(request, 'Produit créé avec succès')
            return redirect('core:product_list')
    else:
        form = ProductForm()
    return render(request, 'core/product_form.html', {'form': form})

@user_passes_test(lambda u: u.is_admin_user())
@login_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produit mis à jour avec succès')
            return redirect('core:product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'core/product_form.html', {
        'form': form,
        'product': product
    })

@user_passes_test(lambda u: u.is_admin_user())
@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Produit supprimé avec succès')
        return redirect('core:product_list')
    return render(request, 'core/product_confirm_delete.html', {'product': product})

@login_required
def product_list(request):
    products = Product.objects.all()
    return render(request, 'core/product_list.html', {'products': products})

@user_passes_test(lambda u: u.is_admin_user())
@login_required
def toggle_admin(request, user_id):
    user = get_object_or_404(User, id=user_id)
    is_admin_before = user.is_admin_user()
    user.is_admin = not is_admin_before
    user.is_staff = user.is_admin
    user.is_superuser = user.is_admin
    user.save()
    if is_admin_before:
        messages.success(request, f"{user.username} est maintenant utilisateur simple")
    else:
        messages.success(request, f"{user.username} est maintenant administrateur")
    return redirect('core:user_management')

@user_passes_test(lambda u: u.is_admin_user())
@login_required
def user_management(request):
    users = User.objects.all()
    return render(request, 'core/user_management.html', {'users': users})

@login_required
def stock_movements(request):
    # Get all movements with related orders
    movements = StockMovement.objects.all().order_by('-created_at')
    
    if request.method == 'POST':
        form = StockMovementForm(request.POST)
        if form.is_valid():
            movement = form.save(commit=False)
            product = movement.product
            
            if movement.movement_type == 'in':
                product.quantity += movement.quantity
            else:
                product.quantity -= movement.quantity
            
            product.save()
            movement.save()
            
            messages.success(request, 'Mouvement de stock enregistré avec succès')
            return redirect('core:stock_movements')
    else:
        form = StockMovementForm()
    
    # Calculate statistics
    total_orders = Order.objects.count()
    total_items = OrderItem.objects.aggregate(total=Sum('quantity'))['total'] or 0
    total_revenue = OrderItem.objects.aggregate(
        total=Sum(F('quantity') * F('price'))
    )['total'] or 0
    total_movements = movements.count()
    
    # Group movements by order
    order_movements = {}
    independent_movements = []
    
    for movement in movements:
        if movement.order:
            order_id = movement.order.id
            if order_id not in order_movements:
                order_movements[order_id] = {
                    'order': movement.order,
                    'movements': []
                }
            order_movements[order_id]['movements'].append(movement)
        else:
            independent_movements.append(movement)
    
    return render(request, 'core/stock_movements.html', {
        'order_movements': order_movements,
        'independent_movements': independent_movements,
        'form': form,
        'total_orders': total_orders,
        'total_items': total_items,
        'total_revenue': total_revenue,
        'total_movements': total_movements
    })


@login_required
def dashboard(request):
    # Statistics
    total_products = Product.objects.count()
    total_orders = Order.objects.count()
    
    if request.user.is_admin:
        user_orders = Order.objects.all()
        recent_orders = Order.objects.select_related('client').order_by('-order_date')[:5]
    else:
        user_orders = Order.objects.filter(client=request.user)
        recent_orders = user_orders.order_by('-order_date')[:5]
    
    total_user_orders = user_orders.count()
    
    # Recent products
    recent_products = Product.objects.order_by('-created_at')[:5]
    
    context = {
        'total_products': total_products,
        'total_orders': total_orders,
        'total_user_orders': total_user_orders,
        'recent_orders': recent_orders,
        'recent_products': recent_products,
    }
    
    return render(request, 'core/dashboard.html', context)


@login_required
def product_list(request):
    products = Product.objects.all()
    return render(request, 'core/product_list.html', {
        'products': products
    })


@login_required
def product_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produit ajouté avec succès.')
            return redirect('core:product_list')
    else:
        form = ProductForm()
    
    return render(request, 'core/product_form.html', {'form': form, 'title': 'Ajouter un produit'})


@login_required
def product_edit(request, pk):
    if not request.user.is_admin:
        messages.error(request, 'Accès refusé. Seuls les administrateurs peuvent modifier des produits.')
        return redirect('core:product_list')
    
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produit modifié avec succès.')
            return redirect('core:product_list')
    else:
        form = ProductForm(instance=product)
    
    return render(request, 'core/product_form.html', {'form': form, 'title': 'Modifier le produit', 'product': product})


@login_required
def product_delete(request, pk):
    if not request.user.is_admin:
        messages.error(request, 'Accès refusé. Seuls les administrateurs peuvent supprimer des produits.')
        return redirect('core:product_list')
    
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Produit supprimé avec succès.')
        return redirect('core:product_list')
    
    return render(request, 'core/product_confirm_delete.html', {'product': product})


@login_required
def order_list(request):
    if request.user.is_admin:
        orders = Order.objects.select_related('client').order_by('-order_date')
    else:
        orders = Order.objects.filter(client=request.user).order_by('-order_date')
    
    return render(request, 'core/order_list.html', {'orders': orders})


@login_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    
    # Check permissions
    if not request.user.is_admin and order.client != request.user:
        messages.error(request, 'Accès refusé. Vous ne pouvez voir que vos propres commandes.')
        return redirect('core:order_list')
    
    return render(request, 'core/order_detail.html', {
        'order': order,
        'can_export': True
    })


@login_required
def order_create(request):
    if request.method == 'POST':
        # Récupérer les données du formulaire
        client_name = request.POST.get('client')
        products = request.POST.getlist('products')
        quantities = request.POST.getlist('quantities')
        prices = request.POST.getlist('prices')
        
        # Créer la commande
        order = Order.objects.create(
            client=request.user,
            status='confirmed'  # Set to confirmed immediately
        )
        
        # Créer les articles de commande
        with transaction.atomic():
            for product_id, quantity, price in zip(products, quantities, prices):
                if product_id and quantity and price:
                    product = Product.objects.get(id=product_id)
                    
                    # Créer l'article
                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=int(quantity),
                        price=float(price)
                    )
                    
                    # Stock will be updated by OrderItem.save() method
            
            # Save order again to ensure everything is consistent
            order.save()
            
            messages.success(request, 'Commande créée avec succès.')
            return redirect('core:order_list')
    
    # Récupérer les produits disponibles
    products = Product.objects.filter(quantity__gt=0)
    
    return render(request, 'core/order_create.html', {
        'products': products
    })


@login_required
def export_order_pdf(request, pk):
    order = get_object_or_404(Order, pk=pk)
    
    # Calculate total by multiplying quantity and price for each item
    order_items = order.items.all()
    total = sum(item.quantity * item.price for item in order_items)
    
    # Render the template to PDF
    template = get_template('core/order_pdf.html')
    context = {
        'order': order,
        'order_items': order_items,
        'total': total
    }
    html = template.render(context)
    
    # Create PDF
    pdf = BytesIO()
    pisa.CreatePDF(html, dest=pdf)
    
    # Set up response
    response = HttpResponse(pdf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=commande_{order.id}.pdf'
    
    return response


@login_required
def export_order_excel(request, pk):
    order = get_object_or_404(Order, pk=pk)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Commande {order.id}"
    
    # Add headers
    headers = ['Produit', 'Quantité', 'Prix unitaire', 'Sous-total']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Add order items
    row_num = 2
    total = 0
    for item in order.items.all():
        ws.cell(row=row_num, column=1, value=item.product.name)
        ws.cell(row=row_num, column=2, value=item.quantity)
        ws.cell(row=row_num, column=3, value=str(item.price))
        subtotal = item.quantity * item.price
        ws.cell(row=row_num, column=4, value=str(subtotal))
        total += subtotal
        row_num += 1
    
    # Add total
    ws.cell(row=row_num, column=3, value='Total:')
    ws.cell(row=row_num, column=4, value=str(total))
    
    # Save workbook to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Set up response
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=commande_{order.id}.xlsx'
    
    return response


@login_required
def user_management(request):
    if not request.user.is_admin:
        messages.error(request, 'Accès refusé. Seuls les administrateurs peuvent gérer les utilisateurs.')
        return redirect('core:dashboard')
    
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'core/user_management.html', {'users': users})


@login_required
def toggle_admin(request, user_id):
    if not request.user.is_admin:
        messages.error(request, 'Accès refusé.')
        return redirect('core:dashboard')
    
    user = get_object_or_404(User, id=user_id)
    user.is_admin = not user.is_admin
    user.save()
    
    status = "administrateur" if user.is_admin else "utilisateur simple"
    messages.success(request, f'{user.username} est maintenant {status}.')
    
    return redirect('core:user_management')